from pathlib import Path
from openpyxl import load_workbook, Workbook, worksheet
from typing import Any, Mapping
import yaml
from datetime import datetime

from actuated_minds_crm.models import Contact


def inspect_worksheet(config: dict[str, Any]) -> worksheet:

    wb = load_workbook(config['data_input']['file_name'],
                      read_only=True,
                      keep_vba=True)
    print("Successful load")
    print("Sheets {}".format(wb.sheetnames))
    selected_sheet: int  = wb.sheetnames.index(config['data_input']['contact_names_sheet'])
    contacts_sheet = wb.worksheets[selected_sheet]
    print("Selected Sheet name: {} and number: {}".format(contacts_sheet.title, selected_sheet))
    headers_list = dict()
    index = 1
    for cell in contacts_sheet[1]:
        headers_list[cell.value] = index
        index = index + 1
    print("Columns: {}".format(headers_list))
    values_row_1 = [cell.value for cell in contacts_sheet[2]]
    values_row_2 = [cell.value for cell in contacts_sheet[3]]
    values_row_3 = [cell.value for cell in contacts_sheet[4]]
    print("First 3 rows: {}/n {}/n {}/n".format(values_row_1, values_row_2, values_row_3))
    return contacts_sheet

def import_contact(headers_row: list, contact_row : list, mapping:dict) -> Contact:
    raw_contact = dict(zip(headers_row, contact_row))
    refactored_contact = dict()
    for key, value in raw_contact.items():
        if key in mapping:
            target_key = mapping[key]
            if value is None:
                continue
            if target_key == "name" and (value is None or str(value).strip() == ""):
                return None  # Skip this contact if the name is empty
            elif isinstance(value, str) and value.strip().upper() == "N/A":
                continue
            elif isinstance(value, datetime):
                refactored_contact[target_key] = value.isoformat()
            elif isinstance(value, bool):
                refactored_contact[target_key] = value
            else:
                refactored_contact[target_key] = value
        else:
            continue
    if not refactored_contact or not refactored_contact.get("name"):
        return None
    contact_object = Contact(**refactored_contact)
    return contact_object


def main():
    config_path = Path(__file__).resolve().with_name('config.yml')
    with open(config_path, 'r', encoding='utf-8') as file:
        config = yaml.safe_load(file)

    contacts_sheet = inspect_worksheet(config)
    contact = Contact(
        name="John Smith",
        region="UK",
        organisation="Example Ventures"
    )

    headers_row = [str(cell.value) for cell in contacts_sheet[1]]
    contacts_list = []
    for row_number, contact_row in enumerate(contacts_sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            contact_object = import_contact(headers_row=headers_row, contact_row=contact_row, mapping=config["column_mapping"])
        except Exception as e:
            raise ValueError(f"Error processing row {row_number}: {e    }") from e
        
        if contact_object is not None:
            contacts_list.append(contact_object)
        else:
            break  #Stop processing further, end reached
    print(contacts_list)

if __name__ == "__main__":
    main()
